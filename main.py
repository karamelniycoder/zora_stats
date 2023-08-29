from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from aiohttp import ClientSession
from web3.eth import AsyncEth
from web3 import Web3
from datetime import datetime
from time import sleep
import asyncio
import requests


RPC = 'https://rpc.zora.energy'
PROXY = 'http://log:pass@ip:port'


class Excel:
    def __init__(self, total_len, eth_price):
        self.lock = asyncio.Lock()
        self.eth_price = float(eth_price)
        workbook = Workbook()
        sheet = workbook.active
        self.file_name = f'{total_len}accs_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'

        sheet['A1'] = 'Address'
        sheet['B1'] = 'ETH Balance'
        sheet['C1'] = 'USD Balance'
        sheet['D1'] = 'Total Txs'
        sheet['E1'] = 'Mint/Purchase Txs'
        sheet['F1'] = 'First Tx Time'
        sheet['G1'] = 'Last Tx Time'

        sheet['A1'].font = Font(bold=True)
        sheet['B1'].font = Font(bold=True)
        sheet['C1'].font = Font(bold=True)
        sheet['D1'].font = Font(bold=True)
        sheet['E1'].font = Font(bold=True)
        sheet['F1'].font = Font(bold=True)
        sheet['G1'].font = Font(bold=True)

        sheet.column_dimensions['A'].width = 45
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 9
        sheet.column_dimensions['E'].width = 18
        sheet.column_dimensions['F'].width = 18
        sheet.column_dimensions['G'].width = 18

        workbook.save('results/'+self.file_name)


    async def edit_table(self, acc_info):
        async with self.lock:
            while True:
                try:
                    workbook = load_workbook('results/'+self.file_name)
                    sheet = workbook.active

                    try: usd_balance = round(float(acc_info['balance']) * self.eth_price, 2)
                    except: usd_balance = 'None'

                    valid_info = [
                        acc_info.get('address'),
                        acc_info.get('balance'),
                        usd_balance,
                        acc_info.get('total_txs'),
                        acc_info.get('mint_txs'),
                        acc_info.get('first_tx_time'),
                        acc_info.get('last_tx_time'),
                    ]
                    sheet.append(valid_info)
                    sheet.cell(sheet.max_row, 1).hyperlink = f'https://explorer.zora.energy/address/{acc_info.get("address")}'
                    sheet.cell(sheet.max_row, 1).font = Font(underline=Font.UNDERLINE_SINGLE)

                    workbook.save('results/'+self.file_name)
                    return True
                except PermissionError:
                    print('Cant save excel file, close it ! ! ! !')
                    sleep(3)
                except Exception as err:
                    print(f'Excel | Cant save excel file: {err} | {acc_info.get("address")}')
                    return False


async def check_mints_count(addr):
    total_mints = 0
    first_tx_time = 'None'
    last_tx_time = 'None'
    adv_params = ''
    async with ClientSession() as session:
        try:
            while True:
                url = f'https://explorer.zora.energy/api/v2/addresses/{addr}/transactions{adv_params}'
                r = await session.get(url, proxy=PROXY)
                r_json = await r.json()
                for tx in r_json['items']:
                    try:
                        if tx['decoded_input']['method_call'].split('(')[0] in ['mint', 'purchase']: total_mints += 1
                        if last_tx_time == 'None':
                            last_tx_data = tx['timestamp']
                            last_tx_time = datetime.strptime(last_tx_data.split('.')[0], "%Y-%m-%dT%H:%M:%S")
                    except: pass
                if r_json['next_page_params'] != None:
                    adv_params = f'?block_number={r_json["next_page_params"]["block_number"]}&index={r_json["next_page_params"]["index"]}&items_count={r_json["next_page_params"]["items_count"]}'
                else:
                    first_tx_data = r_json['items'][-1]['timestamp']
                    first_tx_time = datetime.strptime(first_tx_data.split('.')[0], "%Y-%m-%dT%H:%M:%S")
                    break
        except Exception as err:
            print(f'[{addr}] parse txs error: {err}')

    return total_mints, last_tx_time, first_tx_time


async def check_stats(addr, web3, excel, sem):
    await sem.acquire()
    try:
        acc_info = {'address': addr}
        addr = web3.to_checksum_address(addr)

        balance = await web3.eth.get_balance(addr)
        acc_info['balance'] = round(balance / 10 ** 18, 5)

        nonce = await web3.eth.get_transaction_count(addr)
        acc_info['total_txs'] = nonce

        total_mint_count, last_tx_time, first_tx_time = await check_mints_count(addr)
        acc_info['mint_txs'] = total_mint_count
        acc_info['last_tx_time'] = last_tx_time
        acc_info['first_tx_time'] = first_tx_time

        print(f'[{addr}] balance: {round(balance / 10 ** 18, 5)} ETH | total txs: {nonce} | mint txs: {total_mint_count} | first tx: {first_tx_time} | last tx: {last_tx_time}')
    except Exception as err:
        print(f'[{addr}] get balance error: {err}')
    finally:
        await excel.edit_table(acc_info)
        sem.release()


async def runner(addresses, web3, excel, sem):
    tasks = [check_stats(addr, web3, excel, sem) for addr in addresses]
    await asyncio.gather(*tasks)


def get_eth_price():
    r = requests.get('https://api.binance.com/api/v3/ticker/price?symbol=ETHUSDT')
    return r.json()['price']


if __name__ == '__main__':
    print('''
        ___    ___     ___   __  __    ___    _  _    _____  
       | _ \  | _ \   / _ \ |  \/  |  |_ _|  | \| |  |_   _| 
       |  _/  |   /  | (_) || |\/| |   | |   | .` |    | |     ____
      _|_|_   |_|_\   \___/ |_|__|_|  |___|  |_|\_|   _|_|_    |DD|____T_
    _| """ |_|"""""|_|"""""|_|"""""|_|"""""|_|"""""|_|"""""|   |_ |_____|<  
    "`-0-0-'"`-0-0-'"`-0-0-'"`-0-0-'"`-0-0-'"`-0-0-'"`-0-0-'    @ @-@-@-oo

                https://t.me/ProMintChannel
    ''')

    with open('addresses.txt') as f:
        addresses = f.read().splitlines()

    web3 = Web3(Web3.AsyncHTTPProvider(RPC, request_kwargs={"proxy": PROXY}),modules={"eth": (AsyncEth,)}, middlewares=[])
    old_time = datetime.now()
    max_threads = int(input(f'Loaded {len(addresses)} accs.\nEnter max threads count: '))
    sem = asyncio.Semaphore(max_threads)
    excel = Excel(len(addresses), get_eth_price())

    asyncio.run(runner(addresses, web3, excel, sem))

    new_time = datetime.now()
    print(f'\nsaved in `{excel.file_name}`\ntime spent: {new_time - old_time}')

